import re
import urllib.error
import urllib.request
import requests
import xlwt
from bs4 import BeautifulSoup
import lxml
import csv
import pandas
import numpy
from openpyxl import load_workbook
import time
import random
from fake_useragent import UserAgent

'''
1.从论坛首页抓取详情页网址
2.解析详情页获取评论
3.保存评论到csv
'''


# 从ip池取ip
def get_proxy():
    '''
    需先运行 radis和 ProxyPool.py
    :return:随机ip
    '''
    PROXY_POOL_URL = 'http://localhost:5555/random'
    try:
        response = requests.get(PROXY_POOL_URL)
        if response.status_code == 200:
            return response.text
    except ConnectionError:
        return None


def get_url_m(url, headers, proxy):
    '''
    :param url: 网站首页
    :param headers: 请求头
    :param proxy: 携带IP地址
    :return: 详情页地址(列表)
    '''
    proxies = {
        'http': f'http//{proxy}'
    }
    strhtml = requests.get(url, headers=headers, proxies=proxies)
    soup = BeautifulSoup(strhtml.text, 'lxml')
    info = soup.select('.media-body')
    A, B = 'href="', '">'
    url_l = re.findall(f"{A}.+?{B}", str(info))
    url_m = []
    del url_l[0:2]
    for i in range(len(url_l)):
        j = re.sub('href="', '', str(url_l[i]))
        j = re.sub('">', '', j)
        url_m.append(j)
    # url_m = re.sub((f"{A}|{B}", "", str(url_l)))
    return url_m


# print(get_url_m(url,headers))


def get_talk(url_m, headers, proxy):
    '''
    :param url_m: 网址后缀
    :param headers: 请求头
    :param proxy: 携带IP地址
    :return: 评论（list）
    '''
    proxies = {
        'http': f'http//{proxy}'
    }
    url_s = 'https://bbs.oldmantvg.net/' + str(url_m)
    print(f'携带ip: {proxy} 访问: {url_s}')
    strhtml = requests.get(url_s, headers=headers, proxies=proxies)
    soup = BeautifulSoup(strhtml.text, 'lxml')
    info = soup.select('.media-body')
    # print(info)
    A, B = r'<div ', r'</div>'
    temp = re.findall(f"{A}.+?{B}", str(info), re.S)
    result = re.sub(r'\\t', '', str(temp))
    result = re.sub(r'\\n', '', str(result))
    # print(result)
    C, D, E = r'<div class="message mt-1 break-all">', r'</div>', r'\\xa0'
    talk = re.findall(f"{C}.+?{D}", str(result), re.S)
    talk0 = re.sub(f'{C}', '', str(talk))
    talk0 = re.sub(f'{D}', '', str(talk0))
    talk0 = re.sub(f'{E}', ',', str(talk0))
    # p = re.findall(f'<blockquote class="blockquote">.+?</blockquote>',str(talk0))
    p = re.compile(r'<blockquote class="blockquote">.+?</blockquote>', re.S)
    talk0 = re.sub(p, '', talk0)
    q = re.compile(r'<img.+?/>', re.S)
    talk0 = re.sub(q, '', talk0)
    # print("p:", p)

    # for i in p:
    #     j = str(i)
    #     talk0 = re.sub(j,'***', str(talk0))
    #     print('talk0',talk0)
    # print(talk)
    talk0 = re.sub(r'\\', '', str(talk0))
    talk0 = re.sub(r'<br/>', ',', str(talk0))
    talk0 = re.sub(r'<p>', ',', str(talk0))
    talk0 = re.sub(r'</p>', ',', str(talk0))
    talk0 = re.sub(r"', '", "'***'", str(talk0[1:-1]))
    # print(talk0)

    talklist = talk0.split("***")
    # for i in talk0:
    #     print(i)
    # print("talklist", type(talklist), talklist)
    return talklist


def write_xls(textlist):
    wb = load_workbook('oldmantalk.xlsx')
    sheet = wb.active
    sheet['a1'] = 'text1'
    # 向excel中写入对应的value
    k = 2
    for i in range(len(textlist)):
        sheet.cell(row=k, column=1).value = textlist[i]
        k += 1

    wb.save('oldmantalk.xlsx')
    print(len(textlist), '条数据写入成功！')


# get_talk('thread-57088.htm', headers)
ua = UserAgent()
url_m_list = []
for i in range(35):
    headers = {
        'User-Agent': ua.random}
    i += 1
    url = f'https://bbs.oldmantvg.net/index-{i}.htm'
    url_m = get_url_m(url, headers, get_proxy())  # url_m.type:list
    url_m_list.extend(url_m)
    print(f"第{i}页解析完成")
    # time.sleep(random.randint(1,2))
# print(url_m_list)
all_list = []
for j in range(len(url_m_list)):
    headers = {
        'User-Agent': ua.random}
    print("headers:", headers)
    talk_list = get_talk(url_m_list[j], headers, get_proxy())
    all_list.append(r'\t')
    all_list.extend(talk_list)
    # time.sleep(random.randint(1,3))
write_xls(all_list)
