import re
import urllib.error
import urllib.request
import requests
import xlwt
from bs4 import BeautifulSoup
import lxml
import csv
import pandas
import numpy


def scrapy(url, headers):
    strhtml = requests.get(url, headers=headers)
    soup = BeautifulSoup(strhtml.text, 'lxml')
    info = soup.select('.media-body')
    return str(info)


def cut(info):
    A, B, C, D, E, F = 'htm">', '</a>', '<span class="huux_thread_hlight_style1">', '<span class="huux_thread_hlight_style2">', '<span class="huux_thread_hlight_style3">', '</span>'
    temp = re.findall(f"{A}.+?{B}", str(info))
    result = re.sub(f"{A}|{B}", "", str(temp))
    text = re.sub(f"{C}|{D}|{E}|{F}", "", result)
    print('cutType is ',type(text))
    return text


headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36'}

textlist = []

for i in range(2):
    i += 1
    url = f"https://bbs.oldmantvg.net/index-{i}.htm"
    info = scrapy(str(url), headers)
    tittlelist = cut(info)
    # print(tittlelist)
    tittlelist = tittlelist.strip('[')
    tittlelist = tittlelist.strip(']')
    tittlelist = tittlelist.split(",")
    # print(type(tittlelist), tittlelist)
    textlist.append(tittlelist)

with open(r'oldmantvg.txt', 'w') as f:
    for i in textlist:
        f.write(str(i) + '\n')

# print(type(tittlelist))
# print(len(tittlelist), len(textlist))
# print(len(tittlelist) * len(textlist))

from openpyxl import load_workbook

wb = load_workbook('oldmantvg.xlsx')
sheet = wb.active

sheet['a1'] = '标题'

# 向excel中写入对应的value
k = 2
for i in range(len(textlist)):
    for j in range(len(tittlelist)):
        sheet.cell(row=k, column=1).value = textlist[i][j]
        k += 1

wb.save('oldmantvg.xlsx')
print(len(tittlelist) * len(textlist), '条数据写入成功！')

# f = open('oldmantvg.csv', 'w', encoding='utf-8')
#
# csv_writer = csv.writer(f)
#
# csv_writer.writerow(["标题"])
#
# for k in range(len(tittlelist)):
#     for l in range(len(textlist)):
#         csv_writer.writerow(str([tittlelist[k][l]]))
#
# f.close()
